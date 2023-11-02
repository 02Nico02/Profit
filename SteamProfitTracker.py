import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import time
import math
from datetime import datetime, timedelta, date
import requests
import openpyxl
from steam import Steam
from bs4 import BeautifulSoup
from threading import Thread, Lock
import re
import csv
from urllib.parse import unquote
from openpyxl import load_workbook
from requests.exceptions import RequestException
import tensorflow as tf
import numpy as np
from queue import Queue, Empty
from ratelimit import limits, sleep_and_retry
import sqlite3


def es_numero_real(P):
    try:
        float(P)
        return True
    except ValueError:
        return False


class Benefit_Finder:
    def __init__(self, window):
        self.MAX_REQUESTS = 85
        self.REQUEST_INTERVAL = 60 / self.MAX_REQUESTS
        self.steam_id = ""
        self.nameTxtCookies = "cookiesSteam.txt"
        self.sessionid = ""
        self.steamLoginSecure = ""
        self.api_key = "7B29A9649F50BD169491CBEEDF3775DB"
        self.connect_window_open = False
        self.last_request_time = time.time()
        self.fecha_actual = datetime.now().strftime("%d-%m-%Y")
        self.looking_for_discounts = True  # para buscar o no juegos en descuentos
        self.recopilarInformacion = []
        self.PRECIO_MINIMO_STEAM = 41
        self.PRECIO_MINIMO_SIN_DESCUENTO = self.PRECIO_MINIMO_STEAM / 0.75
        self.modelo_cargado = tf.keras.models.load_model("modelo_entrenado.h5")
        self.urls_packs = []
        self.cola_games = Queue()
        self.cola_guardar_errores_en_excel = Queue()
        self.cola_guardarOferta_y_presentar_juego = Queue()
        self.cola_href_descartados_en_excel = Queue()
        self.cola_guardar_href_en_excel = Queue()
        self.cola_borrar_elemento_en_excel = Queue()
        self.cola_recopilarInformacion = Queue()
        self.cola_appids_errors = Queue()
        self.cola_urls_packs = Queue()
        self.cola_agregar_cant_cromos_juego = Queue()
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36"
        }
        self.hilos = []

        # with open(self.nameTxtCookies, "r") as archivo:
        #     lineas = archivo.readlines()

        # if len(lineas) >= 2:
        #     self.sessionid = lineas[0].strip()
        #     self.steamLoginSecure = lineas[1].strip()
        #     self.steam_id = self.steamLoginSecure[:17]
        # archivo.close()

        self.db_conn = sqlite3.connect("database.db")
        self.create_tables()

        # Obtener los datos del usuario con ID 1
        cursor = self.db_conn.cursor()
        cursor.execute("SELECT sessionid, steamLoginSecure FROM Usuario LIMIT 1")
        usuario_data = cursor.fetchone()

        self.s = requests.Session()

        if usuario_data:
            self.sessionid, self.steamLoginSecure = usuario_data
            if len(self.steamLoginSecure) >= 17:
                self.steam_id = self.steamLoginSecure[:17]
            else:
                self.steam_id = ""
        else:
            # Si el usuario no existe, crea uno con valores nulos
            cursor.execute(
                "INSERT INTO Usuario (id, sessionid, steamLoginSecure) VALUES (?, '', '')",
                (1,),
            )
            self.db_conn.commit()
            self.sessionid = ""
            self.steamLoginSecure = ""
            self.steam_id = ""

        self.s.cookies.update(
            {
                "sessionid": self.sessionid,
                "steamLoginSecure": self.steamLoginSecure,
                "birthtime": "756100801",
                "lastagecheckage": "17-9-1994",
            }
        )

        self.wind = window
        self.wind.title("Buscador de beneficios")
        self.wind.protocol("WM_DELETE_WINDOW", self.on_closing)
        # self.wind.geometry("350x260")  # Aumenta el tamaño de la ventana

        # Contenedor principal
        main_frame = tk.Frame(self.wind)
        main_frame.pack(pady=20)

        self.corroborar_session()

        # Etiqueta de estado de conexión
        self.connected_label = tk.Label(
            main_frame,
            text="Estado: Conectado" if self.connected else "Estado: Desconectado",
            fg="green" if self.connected else "red",
        )
        self.connected_label.grid(
            row=0, column=0, columnspan=2, padx=10, pady=10, sticky="w"
        )

        # Botón de conexión
        self.connect_button = tk.Button(
            main_frame, text="Conectar", command=self.conectar
        )
        self.connect_button.grid(row=0, column=2, padx=10, pady=10, sticky="e")
        if self.connected:
            self.connect_button.config(state=tk.DISABLED)

        # Etiquetas y entradas para precio mínimo y máximo
        tk.Label(main_frame, text="Precio mínimo:").grid(
            row=1, column=0, padx=10, pady=5, sticky="e"
        )
        self.price_min = tk.Entry(main_frame, validate="key")
        self.price_min.grid(row=1, column=1, padx=10, pady=5)
        self.price_min["validatecommand"] = (
            self.price_min.register(es_numero_real),
            "%P",
        )

        tk.Label(main_frame, text="Precio máximo:").grid(
            row=2, column=0, padx=10, pady=5, sticky="e"
        )
        self.price_max = tk.Entry(main_frame, validate="key")
        self.price_max.grid(row=2, column=1, padx=10, pady=5)
        self.price_max["validatecommand"] = (
            self.price_max.register(es_numero_real),
            "%P",
        )

        # self.archivo_salida = "steam_market_prices2.xlsx"
        # print("por abrir el archivo")
        # while True:
        #     try:
        #         self.libro_de_trabajo = openpyxl.load_workbook(self.archivo_salida)
        #         break
        #     except PermissionError:
        #         messagebox.showerror(
        #             "Error",
        #             f"Por favor cierre el archivo Excel {self.archivo_salida} y presione Enter para continuar…",
        #         )
        # print("archivo abierto")

        # Botón de búsqueda
        self.buscar_button = tk.Button(
            main_frame, text="Buscar", command=self.realizar_busqueda
        )
        self.buscar_button.grid(row=3, column=0, pady=10)
        if not self.connected:
            self.buscar_button.config(state=tk.DISABLED)

        # Botón de cancelar
        self.cancelar_button = tk.Button(
            main_frame,
            text="Cancelar",
            command=self.cancelar_busqueda,
            state=tk.DISABLED,
        )
        self.cancelar_button.grid(row=3, column=1, padx=10, pady=10, sticky="e")
        self.busqueda_canelada = False

        # Crear la tabla para mostrar las ofertas agregadas
        self.offer_table = ttk.Treeview(
            self.wind, columns=("appid", "Juego", "Precio", "Descuento", "Ganancia")
        )

        self.offer_table.heading("#0", text="appid")
        self.offer_table.heading("#1", text="Juego")
        self.offer_table.heading("#2", text="Precio", anchor=tk.CENTER)
        self.offer_table.heading("#3", text="Descuento", anchor=tk.CENTER)
        self.offer_table.heading("#4", text="Ganancia", anchor=tk.CENTER)
        self.offer_table.heading("#5", text="% Ganancia", anchor=tk.CENTER)
        self.offer_table.column("#0", width=70)
        self.offer_table.column("#1", width=200)
        self.offer_table.column("#2", width=60)
        self.offer_table.column("#3", width=60)
        self.offer_table.column("#4", width=60)
        self.offer_table.column("#5", width=80)
        self.offer_table.pack()

        # Barra de progreso
        self.progress_frame = tk.Frame(self.wind)
        self.progress_frame.pack()
        self.progress_bar = ttk.Progressbar(
            self.progress_frame, length=300, value=0, mode="determinate"
        )
        self.progress_bar.grid(row=0, column=0, padx=10)
        self.progress_label = tk.Label(self.progress_frame, text="", padx=10)
        self.progress_label.grid(row=0, column=1, sticky="w")

    def on_closing(self):
        if self.busqueda_canelada:
            # Si la búsqueda ya se canceló, puedes cerrar la ventana de inmediato
            self.wind.destroy()
        else:
            # Si la búsqueda está en curso, primero intenta cancelarla
            self.cancelar_busqueda()
            # Luego, verifica si la búsqueda se canceló correctamente
            if self.wait_for_threads_to_finish():
                self.db_conn.close()
                self.wind.destroy()

    def create_tables(self):
        cursor = self.db_conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Usuario (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sessionid TEXT,
                steamLoginSecure TEXT
            )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Juego (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                appid INTEGER UNIQUE,
                nombre TEXT,
                cantCromos INTEGER,
                url TEXT,
                user_id INTEGER,
                FOREIGN KEY (user_id) REFERENCES Usuario(id)
            )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Historial_precio_juego (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                precio REAL,
                descuento INTEGER,
                ganancia REAL,
                fecha DATETIME,
                juego_id INTEGER,
                FOREIGN KEY (juego_id) REFERENCES Juego(id)
            )
        """
        )

        cursor.execute(
            """
        CREATE TABLE IF NOT EXISTS Decision (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            compro BOOLEAN,
            fecha DATE,
            Historial_precio_juego_id INTEGER,
            FOREIGN KEY (Historial_precio_juego_id) REFERENCES Historial_precio_juego(id)
        )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Error (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                cantidad INTEGER,
                ultima_fecha DATE,
                juego_id INTEGER,
                FOREIGN KEY (juego_id) REFERENCES Juego(id)
            )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Cromo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                tipo TEXT,
                url TEXT,
                juego_id INTEGER,
                FOREIGN KEY (juego_id) REFERENCES Juego(id)
            )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS Historial_precio_cromo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                precio REAL,
                cantidad INTEGER,
                fecha DATE,
                cromo_id INTEGER,
                FOREIGN KEY (cromo_id) REFERENCES Cromo(id)
            )
        """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS URL (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT
            )
        """
        )

        self.db_conn.commit()

    def wait_for_threads_to_finish(self):
        for hilo in self.hilos:
            hilo.join()

        # Verifica si todos los hilos han terminado
        for hilo in self.hilos:
            if hilo.is_alive():
                return False

        return True

    def cancelar_busqueda(self):
        self.busqueda_canelada = True
        self.cancelar_button["state"] = tk.DISABLED

    def corroborar_session(self):
        response = self.s.get("https://store.steampowered.com/account/", timeout=10)
        if (
            response.status_code == 200
            and response.url == "https://store.steampowered.com/account/"
        ):
            self.connected = True
        else:
            self.connected = False

    def conectar(self):
        # Función para manejar la conexión
        if self.connected:
            return

        # Verificar si la ventana de conexión ya está abierta
        if self.connect_window_open:
            return

        # Abre una ventana emergente para solicitar los nuevos datos de inicio de sesión
        self.mostrar_ventana_inicio_sesion()

    def mostrar_ventana_inicio_sesion(self):
        # Crear una nueva ventana emergente para la entrada de datos
        self.connect_window_open = True  # Marcar que la ventana está abierta
        self.inicio_sesion_wind = tk.Toplevel()
        self.inicio_sesion_wind.title("Inicio de Sesión")
        self.inicio_sesion_wind.geometry("300x150")

        # Crear un frame para agrupar los elementos
        frame = ttk.Frame(self.inicio_sesion_wind, padding=10)
        frame.grid(row=0, column=0)

        # Etiqueta y entrada para sessionid
        ttk.Label(frame, text="Session ID:").grid(row=0, column=0, sticky="w")
        entry_sessionid = ttk.Entry(frame)
        entry_sessionid.grid(row=0, column=1, padx=5, pady=5)
        entry_sessionid.focus()

        # Etiqueta y entrada para steamLoginSecure
        ttk.Label(frame, text="SteamLoginSecure:").grid(row=1, column=0, sticky="w")
        entry_steamLoginSecure = ttk.Entry(frame)
        entry_steamLoginSecure.grid(row=1, column=1, padx=5, pady=5)

        # Label de errores
        self.message_ventana_inicio_sesion = ttk.Label(
            frame, text="", foreground="red", wraplength=300
        )
        self.message_ventana_inicio_sesion.grid(
            row=2, column=0, columnspan=2, pady=(10, 0)
        )

        # Separador
        # separator = ttk.Separator(frame, orient="horizontal")
        # separator.grid(row=3, column=0, columnspan=2, pady=10, sticky="ew")

        def conectar():
            self.sessionid = entry_sessionid.get()
            self.steamLoginSecure = entry_steamLoginSecure.get()
            if len(self.steamLoginSecure) >= 17:
                self.steam_id = self.steamLoginSecure[:17]
            else:
                self.steam_id = ""
            self.s.cookies.update(
                {
                    "sessionid": self.sessionid,
                    "steamLoginSecure": self.steamLoginSecure,
                    "birthtime": "756100801",
                    "lastagecheckage": "17-9-1994",
                }
            )
            self.corroborar_session()
            if self.connected:
                self.inicio_sesion_wind.destroy()  # Cierra la ventana emergente
                self.connected_label.config(text="Estado: Conectado", fg="green")
                self.connect_button.config(state=tk.DISABLED)
                self.buscar_button["state"] = tk.NORMAL
                self.connect_window_open = False  # Marcar que la ventana se ha cerrado
                self.cookiesModif(self.sessionid, self.steamLoginSecure)
            else:
                self.message_ventana_inicio_sesion["text"] = "credenciales no validas"

        # Botón "Conectar"
        conectar_button = ttk.Button(frame, text="Conectar", command=conectar)
        conectar_button.grid(row=3, column=0, columnspan=2, pady=10)

        # Centrar la ventana emergente en la pantalla principal
        self.inicio_sesion_wind.geometry(
            "+%d+%d" % (self.wind.winfo_x() + 50, self.wind.winfo_y() + 50)
        )

        # Bloquear el tamaño de la ventana emergente
        self.inicio_sesion_wind.resizable(False, False)

    def cookiesModif(self, sessionid, steamLoginSecure):
        # Actualiza los atributos del usuario con ID 1 en la base de datos
        cursor = self.db_conn.cursor()
        cursor.execute(
            "UPDATE Usuario SET sessionid = ?, steamLoginSecure = ? WHERE id = 1",
            (sessionid, steamLoginSecure),
        )
        self.db_conn.commit()

    @sleep_and_retry
    @limits(calls=85, period=60)
    def make_request(self, url):
        response = self.s.get(url, headers=self.headers)
        return response

    def realizar_busqueda(self):
        # Verificar que self.price_min y self.price_max no estén vacíos
        if not self.price_min.get() or not self.price_max.get():
            messagebox.showerror(
                "Error", "Por favor, ingrese los valores de precio mínimo y máximo."
            )
            return

        # Obtener los valores de precio mínimo y máximo como números flotantes
        try:
            price_min = float(self.price_min.get())
            price_max = float(self.price_max.get())
        except ValueError:
            messagebox.showerror(
                "Error", "Los valores de precio deben ser números válidos."
            )
            return

        # Verificar que el precio máximo no sea menor que el precio mínimo
        if price_max < price_min:
            messagebox.showerror(
                "Error", "El precio máximo no puede ser menor que el precio mínimo."
            )
            return

        self.appids_errors = []

        steam = Steam(self.api_key)

        try:
            lista_games = steam.users.get_owned_games(self.steam_id)
        except:
            messagebox.showerror("Error", "Error con la sesion.")
            return

        # # archivo de errores y ofertas
        # self.nameExcel = "datosCromos.xlsx"
        # while True:
        #     try:
        #         self.workbook = openpyxl.load_workbook(self.nameExcel)
        #         break
        #     except PermissionError:
        #         messagebox.showerror(
        #             "Error",
        #             f"Por favor cierre el archivo Excel {self.nameExcel} y presione Enter para continuar…",
        #         )

        # appid de juegos que tengo
        existing_appids = [int(game["appid"]) for game in lista_games["games"]]
        self.juegos_del_usuario = existing_appids
        (
            appid_errors,
            self.posibles_error_cromos,
            self.urls,
        ) = self.get_appids_from_database()

        # Convertir a conjuntos y unir sin repetir valores
        new_set = set(existing_appids) | set(appid_errors)

        # Convertir el conjunto resultante a una lista
        existing_appids = list(new_set)

        # Inicializa la cola de datos
        self.termino = False
        self.termino2 = False
        self.termino3 = False
        self.termino4 = False
        self.termino5 = False
        self.termino6 = False

        self.hilos = []

        self.start_time = None

        hilo_1 = Thread(
            target=self.start_searching_for_Steam_games,
            args=(price_max, price_min, existing_appids),
        )
        hilo_2 = Thread(target=self.recorrer_lista_games)
        hilo_3 = Thread(target=self.recorrer_packs)
        hilo_4 = Thread(target=self.hilo_guardar_datos_en_exel)
        hilo_5 = Thread(target=self.hilo_guardar_informacion_recopilada)
        hilo_6 = Thread(target=self.recorrer_errores_games)

        self.hilos.append(hilo_1)
        self.hilos.append(hilo_2)
        self.hilos.append(hilo_3)
        self.hilos.append(hilo_4)
        self.hilos.append(hilo_5)
        self.hilos.append(hilo_6)

        hilo_1.start()
        hilo_2.start()
        hilo_3.start()
        hilo_4.start()
        hilo_5.start()
        hilo_6.start()

    def convertir_fecha(self, fecha_str):
        try:
            # Suponiendo que el formato de fecha sea "dd-mm-yyyy"
            return datetime.strptime(fecha_str, "%d-%m-%Y")
        except ValueError:
            return None

    def get_appids_from_database(self):
        # Crear un cursor para ejecutar consultas SQL
        cursor = self.db_conn.cursor()

        # Obtener la fecha actual
        today = date.today()

        # Obtener el ID del primer usuario
        cursor.execute("SELECT id FROM Usuario LIMIT 1")
        user_id = cursor.fetchone()[0]

        # Obtener los juegos del usuario por su ID
        cursor.execute("SELECT appid FROM Juego WHERE user_id = ?", (user_id,))

        # Obtener los appids de los juegos del usuario
        appids = [row[0] for row in cursor.fetchall()]
        appid_posibles_errors_cromos = []

        # Realizar la consulta para obtener los datos de la tabla Error
        cursor.execute(
            "SELECT E.ultima_fecha, E.cantidad, J.appid FROM Error AS E INNER JOIN Juego AS J ON E.juego_id = J.id"
        )

        for row in cursor.fetchall():
            ultima_fecha_str, cantidad, appid = row
            ultima_fecha = datetime.strptime(ultima_fecha_str, "%Y-%m-%d").date()
            date_diff = today - ultima_fecha

            if date_diff < timedelta(days=cantidad):
                appids.append(appid)
            else:
                appid_posibles_errors_cromos.append(appid)

        # Realizar la consulta para obtener las URLs de la tabla URL
        cursor.execute("SELECT url FROM URL")

        urls = [row[0] for row in cursor.fetchall()]

        return appids, appid_posibles_errors_cromos, urls

    # def get_appids_from_excel(self):
    #     # seleccionar la segunda hoja
    #     worksheet = self.workbook.worksheets[1]

    #     today = datetime.now()

    #     # Leer todas las filas de la segunda hoja 2 y agregar los appid a un array
    #     appids = []
    #     appid_posibles_errors_cromos = []
    #     for row in worksheet.iter_rows(min_row=1, max_col=4):
    #         appid = row[0].value
    #         colum_error = row[2].value
    #         colum_cant_error = row[3].value
    #         if appid:
    #             if (
    #                 colum_error == "0 cromos"
    #                 or colum_error == "no se encontro el precio"
    #             ):
    #                 if colum_cant_error is not None:
    #                     date_value = row[1].value
    #                     date_value = self.convertir_fecha(date_value)
    #                     if date_value:
    #                         date_diff = today - date_value
    #                         target_days = colum_cant_error
    #                         if date_diff < timedelta(days=target_days):
    #                             try:
    #                                 appids.append(int(appid))
    #                             except ValueError:
    #                                 print(f"Skipped non-integer value: {appid}")
    #                         else:
    #                             try:
    #                                 appid_posibles_errors_cromos.append(int(appid))
    #                             except ValueError:
    #                                 print(f"2. Skipped non-integer value: {appid}")

    #             else:
    #                 try:
    #                     appids.append(int(appid))
    #                 except ValueError:
    #                     # Handle non-integer values here, you can print a message or skip the row
    #                     print(f"Skipped non-integer value: {appid}")

    #     # seleccionar la primera hoja
    #     worksheet = self.workbook.worksheets[0]
    #     # Leer todas las filas de la hoja 1 y agregar los appid a un array
    #     for row in worksheet.iter_rows(min_row=1, max_col=7):
    #         appid = row[6].value
    #         if appid:
    #             appids.append(int(appid))

    #     # seleccionar la primera hoja
    #     worksheet = self.workbook.worksheets[3]
    #     urls = []
    #     # Leer todas las filas de la hoja 1 y agregar los appid a un array
    #     for row in worksheet.iter_rows(min_row=1, max_col=1):
    #         url = row[0].value
    #         urls.append(url)

    #     # Retornar el array de appids
    #     return appids, appid_posibles_errors_cromos, urls

    def search_first_page(self, precioMinimo, precioMaximo, diccionarioPage, specials):
        """
        Busca la primera página de juegos en Steam que cumple con los requisitos de precio mínimo y máximo.

        Args:
            precioMinimo (float): Precio mínimo deseado para los juegos.
            precioMaximo (float): Precio máximo deseado para los juegos.
            diccionarioPage (dict): Diccionario para almacenar los resultados de la búsqueda por página.
            specials (int): Indicador especial para la búsqueda en Steam (puede ser 0 o 1).

        Returns:
            tuple: Una tupla que contiene un indicador de éxito (True o False) y el índice de la página encontrada (si se encuentra).

        """
        start_low = 0
        start_high = 7000

        primeroConPrecioMinimo = -1
        while not self.busqueda_canelada and start_low < start_high:
            start_mid = ((start_low + start_high) // 100) * 50

            current_url = f"https://store.steampowered.com/search/?sort_by=Price_ASC&category1=998%2C10&hidef2p=1&category2=29&specials={specials}&ndl=1&start={start_mid}"

            response = self.make_request(current_url)
            soup = BeautifulSoup(response.content, "html.parser")

            games = soup.find_all(
                "a",
                class_=[
                    "search_result_row",
                    "ds_collapse_flag",
                    "ds_flagged",
                    "ds_owned",
                    "app_impression_tracked",
                ],
            )

            if not games:
                if start_low < start_high:
                    start_high = start_mid
                    continue
                else:
                    self.busqueda_cancelada = True
                    print("error")
                    break

            primer_elemento = games[0]
            ultimo_elemento = games[-1]

            primer_price_element = primer_elemento.find(
                "div", class_="discount_final_price"
            )
            ultimo_price_element = ultimo_elemento.find(
                "div", class_="discount_final_price"
            )

            if not primer_price_element and not ultimo_price_element:
                if start_low < start_high:
                    start_low += 50
                    print(f"sin precios {start_mid}")
                    continue
                else:
                    self.busqueda_cancelada = True
                    print("error 2")
                    break

            if not primer_price_element or not ultimo_price_element:
                if ultimo_price_element:
                    # Buscar el primer elemento dentro de games que tiene "div" con class="discount_final_price"
                    primer_elemento_con_precio = next(
                        (
                            elemento
                            for elemento in games
                            if elemento.find("div", class_="discount_final_price")
                        ),
                        None,
                    )

                    if primer_elemento_con_precio:
                        primer_price_element = primer_elemento_con_precio.find(
                            "div", class_="discount_final_price"
                        )
                        print(primer_price_element)
                    else:
                        # Si no se encuentra ningún elemento con "div" y class="discount_final_price", continúa con la búsqueda
                        self.busqueda_cancelada = True
                        print("error 3")
                        break
                else:
                    self.busqueda_cancelada = True
                    print("error 4")
                    break

            primer_price = self.parse_price(primer_price_element.text.strip())
            ultimo_price = self.parse_price(ultimo_price_element.text.strip())

            if primer_price and primer_price >= precioMinimo:
                primeroConPrecioMinimo = start_mid
                start_high = start_mid
                if ultimo_price and ultimo_price <= precioMaximo:
                    diccionarioPage[start_mid] = games
                continue

            if ultimo_price and ultimo_price >= precioMinimo:
                diccionarioPage[start_mid] = games
                return True, start_mid
            else:
                print(f"primer precio: {primer_price}")
                print(primer_price_element)
                print(f"ultimo precio: {ultimo_price}")
                print(ultimo_price_element)
                print(f"el verdadero o falso: {ultimo_price >= precioMinimo}")
                print(f"{start_low} - {start_high}")
                start_low = start_mid
                continue

        if primeroConPrecioMinimo != -1:
            return True, primeroConPrecioMinimo
        return False, 0

    def tiene_cromos(self, data):
        # Buscar todos los elementos <a> con la clase "game_area_details_specs_ctn"
        game_specs_links = data.find_all("a", class_="game_area_details_specs_ctn")

        result = False
        for link in game_specs_links:
            img = link.find("img", src=lambda x: x and "ico_cards.png" in x)
            if img:
                result = True
                break
        return result

    def profit_pack(self, url):
        try:
            response = self.make_request(url)
            html = response.content

            # Analizar el HTML con BeautifulSoup
            soup = BeautifulSoup(html, "html.parser")

            # obtengo el div del precio y nombre del pack
            contenedor_price_pack = soup.find("div", class_="game_area_purchase_game")
            if contenedor_price_pack is None:
                raise Exception(
                    "No se pudo encontrar el contenedor de precio y nombre del pack de "
                    + url
                )

            precio_pack = contenedor_price_pack.find(
                "div",
                class_=[
                    "discount_final_price",
                    "original",
                ],
            )
            if precio_pack is None:
                # game_purchase_price price original
                precio_pack = contenedor_price_pack.find(
                    "div",
                    class_=[
                        "game_purchase_price",
                        "price",
                        "original",
                    ],
                )
                if precio_pack is None:
                    raise Exception("No se pudo encontrar el precio del pack " + url)

            contenidos_del_pack = soup.find_all(
                "div",
                class_=[
                    "tab_item",
                    "tablet_list_item",
                    "app_impression_tracked",
                ],
            )
            precio_pack = self.parse_price(precio_pack.text.strip())
            descuento = self.obtener_porcentaje_descuento(contenedor_price_pack)
            arrayContenidoConCromos = []

            for contenido in contenidos_del_pack:
                a_contenido = contenido.find("a", class_="tab_item_overlay")
                if a_contenido is None:
                    continue  # Saltar iteración si no se encuentra el enlace
                href = a_contenido.get("href")
                appid = int(href.split("/app/")[1].split("/")[0])

                if appid in self.juegos_del_usuario:
                    continue

                try:
                    response = self.make_request(href)
                    response.raise_for_status()
                    data = BeautifulSoup(response.content, "html.parser")
                    tieneCromo = self.tiene_cromos(data)
                    if tieneCromo:
                        arrayContenidoConCromos.append(appid)
                except RequestException as e:
                    print(f"Error al obtener datos de la URL {href}: {e}")

            sumTotalObtenido = 0
            Totalcromos = 0

            if len(arrayContenidoConCromos) == 0:
                self.cola_href_descartados_en_excel.put(url)
                # self.guardar_href_descartados_en_excel(url)
                return

            for appid in arrayContenidoConCromos:
                try:
                    precio_mas_bajo, cant_cromos = self.cant_cromos_and_price_min_cromo(
                        appid=appid
                    )
                    Totalcromos += cant_cromos
                    totalObtenido = self.get_totalObtenido(precio_mas_bajo, cant_cromos)
                    sumTotalObtenido += totalObtenido
                except Exception as e:
                    print(f"Error al procesar la aplicación {appid}: {e}")

            ganancia = round(sumTotalObtenido - precio_pack, 2)
            if ganancia > 0:
                porcentaje_ganancia = (ganancia / precio_pack) * 100
                self.offer_table.insert(
                    "",
                    tk.END,
                    text="Pack",
                    values=(
                        (url),
                        ("$" + str(precio_pack)),
                        str(descuento),
                        ("$" + str(ganancia)),
                        (str(porcentaje_ganancia) + "%"),
                    ),
                )
                # self.cola_guardar_href_en_excel.put(
                #     url,
                #     precio_pack,
                #     descuento,
                #     len(contenidos_del_pack),
                #     len(arrayContenidoConCromos),
                #     ganancia,
                #     porcentaje_ganancia,
                #     Totalcromos,
                # )
            else:
                print(f"{url} no es profit")

        except RequestException as e:
            print(f"Error al hacer la solicitud HTTP: {e} - en la url: {url}")
        except Exception as e:
            print(f"Error general en la ejecución: {e} - en la url: {url}")

    def clear_table(self):
        """
        Limpia la tabla.
        """
        records = self.offer_table.get_children()
        for element in records:
            self.offer_table.delete(element)

    def extract_appid(self, href):
        """
        Extraer el id del juego de una URL.
        Si es un pack saltara error y la excepcion verificara si el pack es profit o no.
        """
        appid = None
        try:
            appid = int(href.split("/app/")[1].split("/")[0])
        except:
            last_slash_index = href.rfind("/")
            result = href[: last_slash_index + 1]
            if not result in self.urls:
                self.cola_urls_packs.put(result)
                # self.urls_packs.append(result)
        return appid

    def search_steam_sales(self, max_price, min_price, existing_appids, specials):
        """
         Realiza una búsqueda en Steam para encontrar juegos en oferta dentro de un rango de precios.

        Args:
            max_price (float): El precio máximo que se permitirá para los juegos en oferta.
            min_price (float): El precio mínimo que se permitirá para los juegos en oferta.
            existing_appids (list): Lista de identificadores de aplicaciones de juegos existentes para evitar duplicados.
            specials (int): Indicador especial para la búsqueda en Steam (puede ser 0 o 1).

        Returns:
            None
        """
        found_overpriced_game = False
        cantidadJuegosSuperiorAlPrecio = 0

        diccionarioPages = {}
        resultado, start = self.search_first_page(
            min_price,
            max_price,
            diccionarioPages,
            specials,
        )

        if resultado:
            while not self.busqueda_canelada and not found_overpriced_game:
                if start in diccionarioPages:
                    games = diccionarioPages[start]
                else:
                    url = f"https://store.steampowered.com/search/?sort_by=Price_ASC&category1=998%2C10&hidef2p=1&category2=29&specials={specials}&ndl=1&start={start}"
                    response = self.make_request(url)
                    soup = BeautifulSoup(response.content, "html.parser")

                    games = soup.find_all(
                        "a",
                        class_=[
                            "search_result_row",
                            "ds_collapse_flag",
                            "ds_flagged",
                            "ds_owned",
                            "app_impression_tracked",
                        ],
                    )

                    if not games:
                        break

                for game in games:
                    href = game.get("href")
                    appid = self.extract_appid(href)
                    if not appid:
                        continue
                    if appid in existing_appids:
                        continue

                    price_element = game.find("div", class_="discount_final_price")
                    if not price_element:
                        self.cola_guardar_errores_en_excel.put(
                            [appid, "no se encontro el precio"]
                        )
                        continue

                    price = self.parse_price(price_element.text.strip())

                    if not price:
                        found_overpriced_game = True
                        self.cola_guardar_errores_en_excel.put(
                            [appid, "no se encontro el precio"]
                        )
                        continue

                    if price < min_price:
                        continue

                    if price > max_price:
                        # Si es el primer juego que se encuentra asi se saltea, por si fue un error.
                        print(f"Precio mayor: {cantidadJuegosSuperiorAlPrecio}")
                        if cantidadJuegosSuperiorAlPrecio < 1:
                            cantidadJuegosSuperiorAlPrecio += 1
                            continue
                        # si se encuentra un juego con precio mayor al max_price, se establece la variable en True y se sale del bucle while
                        found_overpriced_game = True
                        break
                    elif cantidadJuegosSuperiorAlPrecio > 0:
                        print("reinicio la cantidad de precios mayores")
                        cantidadJuegosSuperiorAlPrecio = 0

                    self.cola_games.put([game, price, appid])
                    self.total_games += 1

                    self.update_progress_bar()

                start += 50  # actualizar para obtener la siguiente página de resultados

    def start_searching_for_Steam_games(self, max_price, min_price, existing_appids):
        """
        Busca las ofertas en Steam que estén por debajo del precio máximo dado y por encima del precio minimo
        No incluye juegos que ya están en la lista `existing_appids`.
        """

        self.total_games = 0
        self.processed_Games = 0

        self.clear_table()

        self.progress_label.config(text="Iniciando Proceso…")
        self.busqueda_canelada = False
        self.cancelar_button["state"] = tk.NORMAL
        specials = 1 if self.looking_for_discounts else 0

        if min_price < self.PRECIO_MINIMO_SIN_DESCUENTO:
            if max_price < self.PRECIO_MINIMO_SIN_DESCUENTO:
                self.search_steam_sales(max_price, min_price, existing_appids, 0)
            else:
                self.search_steam_sales(
                    self.PRECIO_MINIMO_SIN_DESCUENTO - 0.01,
                    min_price,
                    existing_appids,
                    0,
                )
                self.search_steam_sales(
                    max_price,
                    self.PRECIO_MINIMO_SIN_DESCUENTO,
                    existing_appids,
                    specials,
                )

        else:
            self.search_steam_sales(max_price, min_price, existing_appids, specials)

        self.termino = True
        print("1 terminado")

    # def guardar_href_descartados_en_excel(self, url):
    #     # seleccionar la tercera hoja
    #     worksheet = self.workbook.worksheets[3]

    #     # Recorrer las celdas de la primera columna para buscar la url
    #     for row in worksheet.iter_rows(min_row=1, max_col=1):
    #         if row[0].value == url:
    #             return

    #     worksheet.append([url])

    #     # Guardar los cambios en el archivo
    #     self.workbook.save(self.nameExcel)

    def guardar_href_descartados_en_db(self, url, cursor, db_conn):
        # Verificar si la URL ya existe en la tabla URL
        cursor.execute("SELECT id FROM URL WHERE url = ?", (url,))
        url_id = cursor.fetchone()

        if not url_id:
            # Si la URL no existe, la insertamos en la tabla URL
            cursor.execute("INSERT INTO URL (url) VALUES (?)", (url,))
            db_conn.commit()

    # def guardar_href_en_excel(
    #     self,
    #     href,
    #     precioPack,
    #     desc,
    #     contenidoPack,
    #     contenidoConCromo,
    #     ganancia,
    #     porcentajeGanancia,
    #     cromos,
    # ):
    #     # seleccionar la tercera hoja
    #     worksheet = self.workbook.worksheets[2]

    #     # Recorrer las celdas de la primera columna para buscar la href
    #     for row in worksheet.iter_rows(min_row=1, max_col=1):
    #         if row[0].value == href:
    #             return

    #     # Agregar la nueva href a la tercera hoja del archivo
    #     nueva_fila = [
    #         href,
    #         precioPack,
    #         desc,
    #         contenidoPack,
    #         contenidoConCromo,
    #         ganancia,
    #         porcentajeGanancia,
    #         cromos,
    #         self.fecha_actual,
    #     ]
    #     worksheet.append(nueva_fila)

    #     # Guardar los cambios en el archivo
    #     self.workbook.save(self.nameExcel)

    def hilo_guardar_datos_en_exel(self):
        self.termino4 = False

        db_conn = sqlite3.connect("database.db")
        cursor = db_conn.cursor()

        while not self.busqueda_canelada:
            if not self.cola_guardarOferta_y_presentar_juego.empty():
                (
                    name,
                    price,
                    appid,
                    discount,
                    result_profit,
                    cards,
                ) = self.cola_guardarOferta_y_presentar_juego.get()
                self.guardarOferta_y_presentar_Juego(
                    name, price, appid, discount, result_profit, cards, cursor, db_conn
                )

            if not self.cola_href_descartados_en_excel.empty():
                url = self.cola_href_descartados_en_excel.get()
                self.guardar_href_descartados_en_db(url, cursor, db_conn)

            # if not self.cola_guardar_href_en_excel.empty():
            #     (
            #         href,
            #         precioPack,
            #         desc,
            #         contenidoPack,
            #         contenidoConCromo,
            #         ganancia,
            #         porcentajeGanancia,
            #         cromos,
            #     ) = self.cola_guardar_href_en_excel.get()
            #     self.guardar_href_en_excel(
            #         href,
            #         precioPack,
            #         desc,
            #         contenidoPack,
            #         contenidoConCromo,
            #         ganancia,
            #         porcentajeGanancia,
            #         cromos,
            #     )

            if not self.cola_guardar_errores_en_excel.empty():
                appid, error = self.cola_guardar_errores_en_excel.get()
                self.guardar_errores_en_db(appid, error, cursor, db_conn)

            if not self.cola_borrar_elemento_en_excel.empty():
                appid = self.cola_borrar_elemento_en_excel.get()
                self.borrar_elemento_en_db(appid, cursor, db_conn)

            if (
                self.termino
                and self.termino2
                and self.termino3
                and self.termino6
                and self.cola_guardarOferta_y_presentar_juego.empty()
                and self.cola_guardar_errores_en_excel.empty()
                # and self.cola_guardar_href_en_excel.empty()
                and self.cola_href_descartados_en_excel.empty()
                and self.cola_borrar_elemento_en_excel.empty()
            ):
                break

        db_conn.close()
        self.termino4 = True
        print("4 terminado")

    def hilo_guardar_informacion_recopilada(self):
        self.termino5 = False

        db_conn = sqlite3.connect("database.db")
        cursor = db_conn.cursor()

        while not self.busqueda_canelada:
            if not self.cola_recopilarInformacion.empty():
                (
                    href,
                    nombre_decodificado,
                    precio,
                    cant_cromo,
                    date,
                    appid,
                ) = self.cola_recopilarInformacion.get()

                # Verificar si existe un juego con el mismo appid en la tabla Juego
                cursor.execute("SELECT id FROM Juego WHERE appid = ?", (appid,))
                juego_id = cursor.fetchone()

                if not juego_id:
                    print(
                        "el juego no existe, no se puede guardar la informacion del cromo."
                    )
                    continue
                else:
                    juego_id = juego_id[0]

                # Buscar o crear el cromo
                cursor.execute(
                    "SELECT id FROM Cromo WHERE nombre = ? AND tipo = 'cromo' AND juego_id = ?",
                    (nombre_decodificado, juego_id),
                )
                cromo_id = cursor.fetchone()
                if not cromo_id:
                    cursor.execute(
                        "INSERT INTO Cromo (nombre, tipo, url, juego_id) VALUES (?, ?, ?, ?)",
                        (nombre_decodificado, "cromo", href, juego_id),
                    )
                    cromo_id = cursor.lastrowid
                else:
                    cromo_id = cromo_id[0]

                # Insertar un nuevo historial de precio del cromo
                cursor.execute(
                    "INSERT INTO Historial_precio_cromo (precio, cantidad, fecha, cromo_id) VALUES (?, ?, ?, ?)",
                    (precio, cant_cromo, date, cromo_id),
                )
                db_conn.commit()

            if not self.cola_agregar_cant_cromos_juego.empty():
                appid, n_cartas = self.cola_agregar_cant_cromos_juego.get()
                # Verificar si existe un juego con el mismo appid en la tabla Juego
                cursor.execute(
                    "SELECT id, cantCromos FROM Juego WHERE appid = ?", (appid,)
                )
                juego_info = cursor.fetchone()

                if not juego_info:
                    print(
                        "el juego no existe, no se puede actualizar o verificar la cantidad de cromos."
                    )
                    continue
                juego_id, cant_cromos_actual = juego_info
                if cant_cromos_actual != n_cartas:
                    # Si la cantidad de cromos es diferente, actualiza la cantidad en la base de datos
                    cursor.execute(
                        "UPDATE Juego SET cantCromos = ? WHERE id = ?",
                        (n_cartas, juego_id),
                    )
                    db_conn.commit()

            if (
                self.termino
                and self.termino2
                and self.termino3
                and self.cola_recopilarInformacion.empty()
                and self.cola_agregar_cant_cromos_juego.empty()
            ):
                break

        db_conn.close()
        self.termino5 = True
        print("5 terminado")

    def guardar_errores_en_db(self, appid, error, cursor, db_conn):
        # Verificar si existe un juego con el mismo appid en la tabla Juego
        cursor.execute("SELECT id FROM Juego WHERE appid = ?", (appid,))
        juego_id = cursor.fetchone()

        if not juego_id:
            print("el juego no existe, no se puede registrar el error.")
            # Si el juego no existe, no podemos registrar el error, por lo que retornamos
            return

        # Verificar si ya existe un error para este juego en la tabla Error
        cursor.execute(
            "SELECT id, nombre, cantidad FROM Error WHERE juego_id = ?", (juego_id[0],)
        )
        existing_error = cursor.fetchone()

        if existing_error:
            if existing_error[1] == error:
                # Si el error es el mismo, incrementamos la cantidad en 1 y actualizamos la fecha
                cursor.execute(
                    "UPDATE Error SET cantidad = cantidad + 1, ultima_fecha = ? WHERE id = ?",
                    (self.fecha_actual, existing_error[0]),
                )
            else:
                # Si el error es diferente, actualizamos el nombre del error, la cantidad y la fecha
                cursor.execute(
                    "UPDATE Error SET nombre = ?, cantidad = 1, ultima_fecha = ? WHERE id = ?",
                    (error, self.fecha_actual, existing_error[0]),
                )
        else:
            # Si no existe un error para este juego, lo creamos
            cursor.execute(
                "INSERT INTO Error (nombre, cantidad, ultima_fecha, juego_id) VALUES (?, 1, ?, ?)",
                (error, self.fecha_actual, juego_id[0]),
            )

        db_conn.commit()

    # def guardar_errores_en_excel(self, appid, error):
    #     # Seleccionar la segunda hoja del archivo
    #     worksheet = self.workbook.worksheets[1]

    #     # Buscar el appid en la primera columna
    #     for row in worksheet.iter_rows(min_row=1, max_col=4):
    #         if row[0].value == appid:
    #             row[1].value = self.fecha_actual
    #             if row[2].value == error:
    #                 # Verificar si la columna 4 (columna D) es un número y luego sumarle 1
    #                 col4_value = row[3].value
    #                 if isinstance(col4_value, (int, float)):
    #                     row[3].value = col4_value + 1
    #             else:
    #                 print(f"el error es distinto, cambia de {row[2].value} a {error}")
    #                 row[2].value = error
    #                 row[3].value = 1

    #             # Guardar los cambios en el archivo
    #             self.workbook.save(self.nameExcel)
    #             return

    #     # Agregar el nuevo appid a la segunda hoja del archivo
    #     nueva_fila = [appid, self.fecha_actual, error, 1]
    #     worksheet.append(nueva_fila)

    #     # Guardar los cambios en el archivo
    #     self.workbook.save(self.nameExcel)

    def borrar_elemento_en_db(self, appid, cursor, db_conn):
        # Verificar si existe un juego con el mismo appid en la tabla Juego
        cursor.execute("SELECT id FROM Juego WHERE appid = ?", (appid,))
        juego_id = cursor.fetchone()

        if not juego_id:
            # Si el juego no existe, no podemos eliminar el error, por lo que retornamos
            return False

        # Verificar si existe un error para este juego en la tabla Error
        cursor.execute("SELECT id FROM Error WHERE juego_id = ?", (juego_id[0],))
        existing_error = cursor.fetchone()

        if existing_error:
            # Si existe un error, lo eliminamos
            cursor.execute("DELETE FROM Error WHERE id = ?", (existing_error[0],))
            db_conn.commit()
            return True  # Elemento encontrado y eliminado

        return False

    # def borrar_elemento_en_excel(self, appid):
    #     # Seleccionar la segunda hoja del archivo
    #     worksheet = self.workbook.worksheets[1]

    #     for row in worksheet.iter_rows(
    #         min_row=2, max_col=1
    #     ):  # Comenzar desde la fila 2 para omitir encabezados
    #         if row[0].value == appid:
    #             worksheet.delete_rows(
    #                 row[0].row
    #             )  # Eliminar la fila que contiene el appid
    #             self.workbook.save(self.nameExcel)  # Guardar los cambios en el archivo
    #             return True  # Elemento encontrado y eliminado

    #     return False  # El elemento no se encontró en la segunda hoja

    def get_totalObtenido(self, precioCromo, cromosAObtener):
        return precioCromo * 0.85 * cromosAObtener

    def parse_price(self, price_str):
        if not price_str:
            return None  # Si la cadena está vacía, devuelve None
        try:
            # Parsea la etiqueta HTML y extrae el contenido de texto
            soup = BeautifulSoup(price_str, "html.parser")
            price_text = soup.get_text()

            price_text = price_text.replace(".", "")

            if "ARS" in price_text:
                price_text = (
                    price_text.split("ARS")[-1]
                    .strip()
                    .replace(",", ".")
                    .replace("$", "")
                )
            else:
                price_text = price_text.replace("$", "").strip().replace(",", ".")

            # Convierte el precio directamente a flotante
            return float(price_text)
        except ValueError:
            return None

    def take_out_comma(self, num):
        return num.replace(",", "")

    def recorrer_errores_games(self):
        self.termino3 = False

        list_appid_procesadas = []
        while not self.busqueda_canelada:
            if not self.cola_appids_errors.empty():
                appid, precio_del_juego, discount = self.cola_appids_errors.get()
                if appid in list_appid_procesadas:
                    self.cola_guardar_errores_en_excel.put([appid, "0 cromos"])
                    continue
                list_appid_procesadas.append(appid)
                result_profit, cards = self.calculate_profit(
                    appid, precio_del_juego, discount
                )
                if result_profit <= 0:
                    if not self.connected:
                        messagebox.showerror("Error", "Sé cerro la sesión.")
                        break
                    continue
                self.cola_guardarOferta_y_presentar_juego.put(
                    ["-", precio_del_juego, appid, discount, result_profit, cards]
                )
            if self.termino and self.termino2 and self.cola_appids_errors.empty():
                break

        self.termino3 = True
        print("6 terminado")

    def recorrer_packs(self):
        self.termino6 = False
        while not self.busqueda_canelada:
            if not self.cola_urls_packs.empty():
                url_pack = self.cola_urls_packs.get()
                self.profit_pack(url_pack)

            if self.termino and self.cola_urls_packs.empty():
                break
        self.termino6 = True
        print("3 terminado")

    def recorrer_lista_games(self):
        self.termino2 = False

        # Bloqueo para garantizar acceso seguro a la cola
        cola_lock = Lock()
        start_time_set = False

        # Función que será ejecutada por los hilos
        def process_queue():
            nonlocal start_time_set

            db_conn = sqlite3.connect("database.db")
            cursor = db_conn.cursor()

            while not self.busqueda_canelada:
                if not self.cola_games.empty():
                    with cola_lock:
                        game, price, appid = self.cola_games.get()
                        if not start_time_set:
                            self.start_time = time.time()
                            start_time_set = True

                    name_element = game.find("span", class_="title")
                    if not name_element:
                        if appid:
                            cursor.execute(
                                "SELECT id FROM Juego WHERE appid = ?", (appid,)
                            )
                            juego_id = cursor.fetchone()
                            if juego_id:
                                juego_id = juego_id[0]
                            else:
                                # Si el juego no existe, créalo con los valores predeterminados
                                cursor.execute(
                                    "INSERT INTO Juego (appid, nombre, cantCromos, url, user_id) VALUES (?, NULL, NULL, ?, NULL)",
                                    (
                                        appid,
                                        f"https://store.steampowered.com/app/{appid}",
                                    ),
                                )
                                juego_id = cursor.lastrowid

                        self.cola_guardar_errores_en_excel.put(
                            [appid, "error nombre juego"]
                        )
                        continue

                    name = name_element.text.strip()

                    cursor.execute(
                        "SELECT id, nombre FROM Juego WHERE appid = ?", (appid,)
                    )
                    juego_info = cursor.fetchone()
                    if juego_info:
                        juego_id, existing_name = juego_info
                        if existing_name != name:
                            # El nombre es diferente, actualiza el nombre en la base de datos
                            cursor.execute(
                                "UPDATE Juego SET nombre = ? WHERE id = ?",
                                (name, juego_id),
                            )
                            db_conn.commit()
                    else:
                        # Si el juego no existe, créalo con los valores predeterminados
                        cursor.execute(
                            "INSERT INTO Juego (appid, nombre, cantCromos, url, user_id) VALUES (?, ?, NULL, ?, NULL)",
                            (
                                appid,
                                name,
                                f"https://store.steampowered.com/app/{appid}",
                            ),
                        )
                        juego_id = cursor.lastrowid

                    discount_str = self.obtener_porcentaje_descuento(game)
                    discount = int(discount_str.strip("-").strip("%"))

                    # Insertar un nuevo registro en la tabla Historial_precio_juego
                    cursor.execute(
                        "INSERT INTO Historial_precio_juego (precio, descuento, ganancia, fecha, juego_id) VALUES (?, ?, ?, ?, ?)",
                        (price, discount, None, datetime.now(), juego_id),
                    )
                    db_conn.commit()

                    self.process_game(name, price, appid, discount)
                if self.termino and self.cola_games.empty():
                    break
            db_conn.close()

        # Crear dos hilos para procesar la cola
        thread1 = Thread(target=process_queue)
        thread2 = Thread(target=process_queue)
        thread3 = Thread(target=process_queue)

        # Iniciar los hilos
        thread1.start()
        thread2.start()
        thread3.start()

        # Esperar a que ambos hilos terminen
        thread1.join()
        thread2.join()
        thread3.join()

        self.termino2 = True
        print("2 terminado")

    def update_progress_bar(self):
        self.progress_bar.configure(maximum=self.total_games)
        self.progress_bar["value"] = self.processed_Games

        if self.start_time is not None:
            # En algún punto del código, calcula el tiempo transcurrido
            elapsed_time = time.time() - self.start_time

            # Convierte el tiempo en un objeto timedelta
            time_delta = timedelta(seconds=elapsed_time)

            # Formatea el tiempo en el formato "HH:MM:SS" sin microsegundos
            time_formatted = str(time_delta).split(".")[0]
            self.progress_label.config(
                text=f"Procesados: {self.processed_Games}/{self.total_games} | Tiempo: {time_formatted}"
            )
        else:
            self.progress_label.config(
                text=f"Procesados: {self.processed_Games}/{self.total_games}"
            )

    def obtener_porcentaje_descuento(self, game):
        discount_element = game.find("div", {"class": "discount_pct"})
        return discount_element.text.strip() if discount_element else "0%"

    def process_game(self, name, price, appid, discount):
        self.last_request_time = time.time()
        result_profit, cards = self.calculate_profit(appid, price, discount)
        self.processed_Games += 1

        self.update_progress_bar()
        if result_profit <= 0:
            return

        self.cola_guardarOferta_y_presentar_juego.put(
            [name, price, appid, discount, result_profit, cards]
        )

    def esperar(self):
        current_time = time.time()
        elapsed_time = current_time - self.last_request_time

        if elapsed_time < self.REQUEST_INTERVAL:
            time.sleep(self.REQUEST_INTERVAL - elapsed_time)

    def fetch_market_page(self, p, appid):
        """
        Realiza una solicitud para obtener una página del mercado de Steam Community.

        Args:
            s (requests.Session): Sesión de solicitudes previamente configurada.
            p (int): Número de página del mercado a obtener.
            appid (int): Identificador de la aplicación del juego en el mercado de Steam.

        Returns:
            BeautifulSoup: Un objeto BeautifulSoup que contiene el análisis del contenido HTML de la página del mercado de Steam.

        """
        url = f"https://steamcommunity.com/market/search?q=&category_753_Game%5B%5D=tag_app_{appid}&category_753_cardborder%5B%5D=tag_cardborder_0&category_753_item_class%5B%5D=tag_item_class_2&appid=753#p{p}_popular_desc"
        self.last_request_time = time.time()
        market_doc2 = self.make_request(url)
        marketsoup2 = BeautifulSoup(market_doc2.content, "html.parser")
        return marketsoup2

    def datos_de_los_cromos(
        self,
        marketsoup2,
        n_errores_precio,
        precio_mas_bajo,
        sumatoria_de_precios,
        appid,
    ):
        num_name_cromo = marketsoup2.find_all("a", class_="market_listing_row_link")
        cromos = marketsoup2.find_all(
            "div",
            class_=[
                "market_listing_row",
                "market_recent_listing_row",
                "market_listing_searchresult",
            ],
        )

        for indice, cromo in enumerate(cromos):
            parts = num_name_cromo[indice]["href"].split("/")
            cromo_name = parts[-1]

            # Obtener el valor del atributo href
            href = num_name_cromo[indice].get("href")

            nombre_decodificado = unquote(cromo_name)

            price_element = cromo.select('span.normal_price:not([class*=" "])')[0]

            # Selecciona el elemento span con la clase "market_listing_num_listings_qty"
            cant_cromo = cromo.find(
                "span", class_="market_listing_num_listings_qty"
            ).text.strip()

            if not price_element:
                n_errores_precio += 1
                continue

            precio = self.parse_price(price_element.text.strip())

            if not precio:
                n_errores_precio += 1
                continue

            sumatoria_de_precios += precio
            precio_mas_bajo = min(precio_mas_bajo, precio)

            self.cola_recopilarInformacion.put(
                [
                    href,
                    nombre_decodificado,
                    precio,
                    int(self.take_out_comma(cant_cromo)),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    appid,
                ]
            )

        return precio_mas_bajo, sumatoria_de_precios, n_errores_precio

    def calculate_profit(self, appid, precio_del_juego, discount):
        """
        Calcula el posible beneficio al comprar y vender cromos de un juego en el mercado de Steam Community.

        Args:
            appid (int): Identificador de la aplicación del juego en el mercado de Steam.
            precio_del_juego (float): El precio del juego en el mercado de Steam.
            discount (str): El descuento aplicado al juego como cadena (puede ser "-" si no hay descuento).

        Returns:
            tuple: Una tupla que contiene el beneficio calculado y la cantidad de cartas a obtener para obtener ese beneficio.

        """
        # Endpoint de la API de Steam para obtener información sobre los cromos de un juego
        p = 1
        marketsoup2 = self.fetch_market_page(p, appid)

        # Lee el total de números de carta
        numeros_de_cartas = marketsoup2.find(id="searchResults_total")

        if not numeros_de_cartas:
            self.corroborar_session()

        n_cartas = (
            int(numeros_de_cartas.get_text().replace(",", ""))
            if numeros_de_cartas
            else 0
        )

        if n_cartas == 0:
            self.cola_appids_errors.put([appid, precio_del_juego, discount])
            return (0, 0)

        if appid in self.posibles_error_cromos:
            self.cola_borrar_elemento_en_excel.put(appid)

        num_page = math.ceil(n_cartas / 10)
        drop_cartas = n_cartas // 2 + (1 if n_cartas % 2 == 1 else 0)
        sumatoria_de_precios = 0
        n_errores_precio = 0
        precio_mas_bajo = float("inf")

        self.cola_agregar_cant_cromos_juego.put([appid, n_cartas])

        (
            precio_mas_bajo,
            sumatoria_de_precios,
            n_errores_precio,
        ) = self.datos_de_los_cromos(
            marketsoup2,
            n_errores_precio,
            precio_mas_bajo,
            sumatoria_de_precios,
            appid,
        )

        if self.get_totalObtenido(precio_mas_bajo, drop_cartas) < precio_del_juego:
            return (0, 0)

        if num_page > 1:
            for p in range(2, num_page + 1):
                marketsoup2 = self.fetch_market_page(p, appid)
                (
                    precio_mas_bajo,
                    sumatoria_de_precios,
                    n_errores_precio,
                ) = self.datos_de_los_cromos(
                    marketsoup2,
                    n_errores_precio,
                    precio_mas_bajo,
                    sumatoria_de_precios,
                    appid,
                )
                if (
                    self.get_totalObtenido(precio_mas_bajo, drop_cartas)
                    < precio_del_juego
                ):
                    return (0, 0)

        if n_errores_precio >= n_cartas:
            self.cola_guardar_errores_en_excel.put([appid, "errores con cromos"])
            # self.guardar_errores_en_excel(appid, "errores con cromos")
            return (0, 0)

        promedio_precios = round(
            self.get_totalObtenido(sumatoria_de_precios, drop_cartas)
            / (n_cartas - n_errores_precio),
            2,
        )
        profit = round(
            self.get_totalObtenido(precio_mas_bajo, drop_cartas) - precio_del_juego, 2
        )
        return (profit, drop_cartas)

    def cant_cromos_and_price_min_cromo(self, appid):
        # Endpoint de la API de Steam para obtener información sobre los cromos de un juego
        p = 1
        marketsoup2 = self.fetch_market_page(p, appid)

        # Lee el total de números de carta
        numeros_de_cartas = marketsoup2.find(id="searchResults_total")

        if not numeros_de_cartas:
            self.corroborar_session()

        n_cartas = (
            int(numeros_de_cartas.get_text().replace(",", ""))
            if numeros_de_cartas
            else 0
        )

        if n_cartas == 0:
            # self.appids_errors.append((appid, precio_del_juego))
            # self.esperar()
            return (0, 0)

        num_page = math.ceil(n_cartas / 10)
        drop_cartas = n_cartas // 2 + (1 if n_cartas % 2 == 1 else 0)
        sumatoria_de_precios = 0
        n_errores_precio = 0
        precio_mas_bajo = float("inf")

        while True:
            cromos = marketsoup2.find_all(
                "div",
                class_=[
                    "market_listing_row",
                    "market_recent_listing_row",
                    "market_listing_searchresult",
                ],
            )

            for indice, cromo in enumerate(cromos):
                price_element = cromo.select('span.normal_price:not([class*=" "])')[0]

                if not price_element:
                    n_errores_precio += 1
                    continue

                precio = self.parse_price(price_element.text.strip())

                if not precio:
                    n_errores_precio += 1
                    continue

                sumatoria_de_precios += precio
                precio_mas_bajo = min(precio_mas_bajo, precio)

            # self.esperar()

            p += 1
            if p > num_page:
                break

            self.last_request_time = time.time()
            marketsoup2 = self.fetch_market_page(p, appid)

        if n_errores_precio >= n_cartas:
            return (0, 0)

        return (precio_mas_bajo, drop_cartas)

    def guardarOferta_y_presentar_Juego(
        self, name, price, appid, discount, result_profit, cards, cursor, db_conn
    ):
        porcentaje_ganancia = round(result_profit * 100 / price, 2)

        # Obtener el ID del juego con el appid dado
        cursor.execute("SELECT id FROM Juego WHERE appid = ?", (appid,))
        juego_id = cursor.fetchone()
        if juego_id:
            juego_id = juego_id[0]
            # Buscar el último historial de precio del juego
            cursor.execute(
                "SELECT id FROM Historial_precio_juego WHERE juego_id = ? ORDER BY fecha DESC LIMIT 1",
                (juego_id,),
            )
            historial_id = cursor.fetchone()

            if historial_id:
                historial_id = historial_id[0]

                # Actualizar el atributo ganancia en el último historial
                cursor.execute(
                    "UPDATE Historial_precio_juego SET ganancia = ? WHERE id = ?",
                    (result_profit, historial_id),
                )
                db_conn.commit()

        # self.guardar_ofertas(
        #     name, price, discount, result_profit, porcentaje_ganancia, appid, cards
        # )

        # Agregar oferta a la tabla
        self.offer_table.insert(
            "",
            tk.END,
            text=appid,
            values=(
                (name),
                ("$" + str(price)),
                str(discount),
                ("$" + str(result_profit)),
                (str(porcentaje_ganancia) + "%"),
            ),
        )

    # def guardar_ofertas(
    #     self, name, price, discount, result_profit, porcentaje_ganancia, appid, cards
    # ):
    #     # Seleccionar la primera hoja del archivo
    #     worksheet = self.workbook.worksheets[0]

    #     # Obtener la última fila en la hoja
    #     last_row = worksheet.max_row + 1

    #     # le quito el % al descuento y lo convierto en numero
    #     if len(discount) > 1:
    #         discount_sin_signo = discount.rstrip("%")
    #         numero_discount = float(discount_sin_signo)
    #         num_para_el_modelo = abs(numero_discount)
    #     else:
    #         numero_discount = discount
    #         num_para_el_modelo = 0

    #     # Viendo que haria el modelo entrenado
    #     nueva_entrada = np.array(
    #         [[price, num_para_el_modelo, result_profit, porcentaje_ganancia]],
    #         dtype=float,
    #     )
    #     resultado = self.modelo_cargado.predict(nueva_entrada)
    #     if resultado > 0.5:
    #         decision = "Compralo"
    #     else:
    #         decision = "No lo Compres"
    #     print(decision)

    #     # Escribir los datos en las celdas correspondientes
    #     nueva_fila = [
    #         name,
    #         price,
    #         numero_discount,
    #         result_profit,
    #         porcentaje_ganancia,
    #         cards,
    #         appid,
    #         self.fecha_actual,
    #         "",
    #         "",
    #         "",
    #         decision,
    #     ]
    #     worksheet.append(nueva_fila)

    #     # Guardar los cambios en el archivo Excel
    #     self.workbook.save(self.nameExcel)


if __name__ == "__main__":
    windows = tk.Tk()
    application = Benefit_Finder(windows)
    windows.mainloop()
